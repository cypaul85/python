import openpyxl as xl

wb = xl.load_workbook('Scores.xlsx')

ws = wb.active

for r in ws.rows:
    row_index = r[0].value
    kor = r[1].value
    eng = r[2].value
    math = r[3].value
    sum = kor + eng + math
    avg = sum / 3

    ws.cell(row=row_index, column=5).value = sum
    ws.cell(row=row_index, column=6).value = avg

    print(kor, eng, math, sum, avg)
wb.save("Scores3.xlsx")
wb.close()
